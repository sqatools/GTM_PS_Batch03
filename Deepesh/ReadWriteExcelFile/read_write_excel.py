import openpyxl

def read_excel_file(filename, row, col):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    cell = sheet.cell(row=row, column=col)
    print(cell.value)

#read_excel_file("test_data.xlsx", row=1, col=1)
#read_excel_file("test_data.xlsx", row=2, col=1)

def read_excel_file_with_sheetname(filename, sheet_name, cell_name):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)


read_excel_file_with_sheetname("test_data.xlsx", sheet_name="Sheet2", cell_name="A1")
read_excel_file_with_sheetname("test_data.xlsx", sheet_name="Sheet2", cell_name="B2")

print("_"*50)
for i in range(1, 5):
    read_excel_file_with_sheetname("test_data.xlsx", sheet_name="Sheet2", cell_name=f"A{i}")
