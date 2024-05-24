from openpyxl import Workbook

wb = Workbook()
wb.save(filename ="test2_data.xlsx")



import openpyxl


wb = openpyxl.load_workbook("test2_data.xlsx")
sheet_obj = wb.active
cell1 = sheet_obj.cell(row = 1, column =1)
cell1.value ="Anu"
cell2 = sheet_obj.cell(row = 2, column =1)
cell2.value ="Ammi"
cell3 = sheet_obj.cell(row = 3, column =1)
cell3.value ="Adhu"
cell4 = sheet_obj.cell(row = 4, column =1)
cell4.value ="monu"

print(cell1.value)
print(cell2.value)
print(cell3.value)
print(cell4.value)







