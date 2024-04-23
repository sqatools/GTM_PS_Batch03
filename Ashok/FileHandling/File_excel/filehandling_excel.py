import openpyxl

"""
openpyxl is module, contain the "load_workbook" function.
1. load workbook 
2. find the sheet 
3. specify the row, column 
4. print the cell value
"""


def read_excel_file1(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    cell_value = sheet.cell(row=1, column=1)
    print(cell_value.value)


read_excel_file1("Book.xlsx")


def read_excel_file2_withparam(filename, row, col):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    cell = sheet.cell(row=row, column=col)
    print(cell.value)


read_excel_file2_withparam('Book.xlsx', 1, 1)
read_excel_file2_withparam('Book.xlsx', 1, 2)
read_excel_file2_withparam('Book.xlsx', 2, 3)


def read_excel_with_sheetname_cellname(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb['Sheet2']
    cell = sheet['A1']
    print(cell.value)


read_excel_with_sheetname_cellname('Book.xlsx')


def read_excel_with_sheetname_cellname_param(filename, sheetname, cell):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    cell = sheet[cell]
    print(cell.value)


read_excel_with_sheetname_cellname_param("Book.xlsx", "Sheet2", "A3")
read_excel_with_sheetname_cellname_param("Book.xlsx", "Sheet1", "A1")


def read_excel_with_sheetname_without_cellname(filename, sheetname, row, col):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    cell = sheet.cell(row=row, column=col)
    print(cell.value)


read_excel_with_sheetname_without_cellname("Book.xlsx", "Sheet2", row=1, col=3)
read_excel_with_sheetname_without_cellname("Book.xlsx", "Sheet2", row=5, col=3)


def read_excel_with_sheet_loops(filename, sheetname, col):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    cell = sheet[col]
    print(cell.value)


for i in range(1, 6):
    read_excel_with_sheet_loops("Book.xlsx", "Sheet1", f"A{i}")


def write_content_to_file_without_overwrite(filename, sheet_name, cell_name, cell_value):
    wb = openpyxl.load_workbook(filename)
    wr = wb[sheet_name]
    cell = wr[cell_name]
    cell.name = cell_value
    wb.save(filename)


write_content_to_file_without_overwrite("Book 1.xlsx", "Sheet1", "A1", "India")
