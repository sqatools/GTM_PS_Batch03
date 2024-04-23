import openpyxl

#1.
# def readExcelFile(filename,row,col):
#     wb=openpyxl.load_workbook(filename)
#     sheet=wb.active
#     cell=sheet.cell(row=row,column=col)
#     print(cell.value)
#
# readExcelFile("SqlData.xlsx",row=2,col=2)         # ANJU THOMAS
# readExcelFile("SqlData.xlsx",row=2,col=3)         # 17

#2.
# def read_excel_file_with_sheetname(filename, sheet_name, cell_name):
#     wb=openpyxl.load_workbook(filename)
#     sheet=wb[sheet_name]
#     cell=sheet[cell_name]
#     print(cell.value)
#
# read_excel_file_with_sheetname("SqlData.xlsx", sheet_name="Sheet1", cell_name="B1")      # NAME
# read_excel_file_with_sheetname("SqlData.xlsx", sheet_name="Sheet1", cell_name="B2")      # ANJU THOMAS


#3.
# def read_excel_file_with_sheetname(filename, sheet_name, cell_name):
#     wb=openpyxl.load_workbook(filename)
#     sheet=wb[sheet_name]
#     cell=sheet[cell_name]
#     print(cell.value)
#
# for i in range(1, 11):                                           # We will get all the data from that particular cell ({B{i}).
#     read_excel_file_with_sheetname("SqlData.xlsx", sheet_name="Sheet1", cell_name=f"B{i}")

#4.
# def read_excel_file_with_sheetname(filename, sheet_name, cell_name):
#     wb=openpyxl.load_workbook(filename)
#     sheet=wb[sheet_name]
#     cell=sheet[cell_name]
#     print(cell.value)
#
# for i in range(1, 11):         # We will get whole data that particular cell(B{i}) and matching value of cell(C{i})
#     read_excel_file_with_sheetname("SqlData.xlsx", sheet_name="Sheet1", cell_name=f"B{i}")
#     read_excel_file_with_sheetname("SqlData.xlsx", sheet_name="Sheet1", cell_name=f"C{i}")
#


#5.Write_content_to_file_without_overwrite :
# def write_excel_file_with_sheetname(filename, sheet_name, cell_name,cell_value):
#     wb = openpyxl.load_workbook(filename)
#     sheet_obj = wb[sheet_name]
#     cell = sheet_obj[cell_name]
#     cell.value = cell_value
#     wb.save(filename)
# write_excel_file_with_sheetname('SqlData.xlsx', sheet_name="Sheet1", cell_name='A11',cell_value='6234')
# write_excel_file_with_sheetname('SqlData.xlsx', sheet_name="Sheet1", cell_name='B11',cell_value='NISHA LESLY')

#6.Do calculation on Excel file:
# def write_excel_file_with_sheetname(filename, sheet_name, cell_name,cell_value):
#     wb = openpyxl.load_workbook(filename)
#     sheet_obj = wb[sheet_name]
#     cell = sheet_obj[cell_name]
#     cell.value = cell_value
#     wb.save(filename)
# write_excel_file_with_sheetname('SqlData.xlsx', sheet_name="Sheet1", cell_name='A12',cell_value='=SUM(A2:A11)')
# write_excel_file_with_sheetname('SqlData.xlsx', sheet_name="Sheet1", cell_name='B11',cell_value='NISHA JOBY')


#7.
# def write_excel_file_with_sheetname(filename, sheet_name, rowNum,colNum,list1):
#     wb = openpyxl.load_workbook(filename)
#     sheet_obj = wb[sheet_name]
#     for row in list1:
#         cell = sheet_obj.cell(rowNum, colNum)
#         cell.value = row
#         rowNum=rowNum+1
#     wb.save(filename)
# list1=['ANU','MEERA','MANU','SNEHA','VIPIN']
# write_excel_file_with_sheetname("SqlData.xlsx", "Sheet1",12, 2, list1)


#8.
# import random
# def write_excel_file_with_sheetname(filename, sheet_name, cell_value,cell_name):
#     wb = openpyxl.load_workbook(filename)
#     sheet_obj = wb[sheet_name]
#     cell = sheet_obj[cell_name]
#     cell.value = cell_value
#     wb.save(filename)
# for i in range(10):
#     rand_num =random.randint(1000, 2000)
#     char = chr(65+i)
#     write_excel_file_with_sheetname("SqlData.xlsx", "Sheet1", f"{char}12", rand_num)